import pandas as pd
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter
import statistics
import numpy
import argparse

parser = argparse.ArgumentParser(description="Strips CT information saved in html file generated by Roche qPCR machines and automatically"
                                             "calculates useful RNAi statistics. For internal use in the Gabbiani/Zong Labs (code not generalized enough for"
                                             "public use).")

parser.add_argument("html_path", type=str, help="The path to the html file (named abs quant001.html by default) that contains the CT values")
parser.add_argument("design_excel", type=str, help="The path to the excel file with the design data for the qPCR. Analysis will be saved to new"
                                                   "sheet in this workbook.")

#Optional argument
parser.add_argument("-p", "--plateout", type=str, help="Used to specify an alternate output path for just the stripped CT values (not the analysis)")

args = vars(parser.parse_args())

htmlPath = args["html_path"]
wb_path = args["design_excel"]

outfile_csv_name =  args["plateout"] if args["plateout"] is not None else 'qPCR_plate.csv'

CONTROL_GENE = "GAPDH"

# Method to extract the CT values from a Roche HTML file
def roche_html_extractor(filepath):
    with open(filepath) as file:
        soup = BeautifulSoup(file, 'html.parser')
    tables = pd.read_html(str(soup))

    df = tables[0]

    position_column_index = None
    qc_column_index = None
    positions = []
    qc_values = []

    df = df.astype(str)

    # Search for "Position" and "Cq" within the DataFrame columns
    for col in df.columns:
        if df[col].str.contains("Position").any():
            position_column_index = col
        if df[col].str.strip().eq("Cq").any():
            qc_column_index = col
        if position_column_index is not None and qc_column_index is not None:
            break  # Stop searching once both columns are found (other Cq's are different things)

    if position_column_index is not None and qc_column_index is not None:
        positions = df[position_column_index].dropna().tolist()
        qc_values = df[qc_column_index].dropna().tolist()

    position_qc_dict = dict(zip(positions, qc_values))

    # Filtering out just to have values A1-H12
    pattern = re.compile(r'^[A-H](?:[1-9]|1[0-2])$')
    filtered_dict = {k: v for k, v in position_qc_dict.items() if pattern.match(k)}

    plate_layout = [["" for _ in range(12)] for _ in range(8)]

    for key, value in filtered_dict.items():
        row, col = key_to_coordinates(key)
        plate_layout[row][col] = float(value)

    return pd.DataFrame(plate_layout, index=[chr(i) for i in range(ord('A'), ord('I'))],
                            columns=[str(i) for i in range(1, 13)])

# Method to apply the design matrix from excel spreadsheet to the extracted CT values to complete the dataset
def apply_design_xlsx(excelfile, qpcr_plate):
    wb = load_workbook(excelfile, data_only=True)
    ws = wb.worksheets[0]

    # Find the "master mix" cell and look below it for gene/color mapping
    legend_start_cell = None
    gene_legend = {}

    # Scan for the "master mix" cell
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "master mix" in cell.value.lower():
                legend_start_cell = (cell.row + 1, cell.column)  # one cell down
                break
        if legend_start_cell:
            break

    # Extract gene names and their associated colors from the legend row
    if legend_start_cell:
        legend_row = legend_start_cell[0]
        for col in range(legend_start_cell[1], ws.max_column + 1):
            cell = ws.cell(row=legend_row, column=col)
            gene_name = cell.value
            fill = cell.fill
            if gene_name and fill.fgColor and fill.fgColor.type == 'rgb':
                rgb = fill.fgColor.rgb[-6:]  # Strip AR part of ARGB
                gene_legend[rgb] = gene_name
            else:
                break

    # Now, we will obtain the samples and create a pandas dataframe
    target_rows = [chr(i) for i in range(ord('A'), ord('I'))]
    target_cols = [str(i) for i in range(1, 13)]

    # Build a DataFrame for sample names from the Excel sheet
    sample_data = []
    gene_data = []

    # openpyxl uses 1-based indexing; A=1, B=2, ..., H=8
    for row_idx, row_letter in enumerate(target_rows, start=1):
        row_values = []
        gene_values = []
        for col_idx in range(1, 13):  # Columns 1 to 12
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            cell_value = cell.value
            fill = cell.fill
            if fill.fgColor and fill.fgColor.type == 'rgb':
                rgb = fill.fgColor.rgb[-6:]
                if rgb in gene_legend.keys():
                    gene_values.append(gene_legend[rgb])
                else:
                    gene_values.append("missing_gene")
            else: gene_values.append("empty")
            row_values.append(cell_value)
        sample_data.append(row_values)
        gene_data.append(gene_values)

    sample_df = pd.DataFrame(sample_data, index=target_rows, columns=target_cols)
    gene_df = pd.DataFrame(gene_data, index=target_rows, columns=target_cols)

    # Melt both dataframes into long format
    ct_long = qpcr_plate.reset_index().melt(id_vars='index', var_name='Column', value_name='CT')
    ct_long.rename(columns={'index': 'Row'}, inplace=True)

    sample_long = sample_df.reset_index().melt(id_vars='index', var_name='Column', value_name='Sample')
    sample_long.rename(columns={'index': 'Row'}, inplace=True)

    gene_long = gene_df.reset_index().melt(id_vars='index', var_name='Column', value_name='Gene')
    gene_long.rename(columns={'index': 'Row'}, inplace=True)

    # Merge
    merged_df = pd.merge(ct_long, sample_long, on=['Row', 'Column'])
    merged_df = pd.merge(merged_df, gene_long, on=['Row', 'Column'])

    # Set MultiIndex
    merged_df.set_index(['Sample', 'Row', 'Column', 'Gene'], inplace=True)

    # Final result
    final_df = merged_df[['CT']]

    print(final_df.head())

    wb.close()
    return final_df

# Mapping function to convert keys to coordinates
def key_to_coordinates(key):
    # Map letter to row index (A -> 0, B -> 1, ..., H -> 7)
    row = ord(key[0]) - ord('A')
    # Convert number to column index (1 -> 0, 2 -> 1, ..., 12 -> 11)
    col = int(key[1:]) - 1
    return (row, col)

def dixons_q_test_n3(values):
    if len(values) != 3:
        raise ValueError("This function only works with exactly 3 values.")

    values = sorted(values)
    Q_crit = 0.941  # 95% confidence for n = 3
    range_ = values[2] - values[0]

    # however, if ct range is less than 1.5 we should ignore this otherwise if the range is really small (like 0.04)
    # but one value is at the extreme end of that range this will return it as an outlier
    if range_ <= 1.5:
        return None

    Q_min = (values[1] - values[0]) / range_
    Q_max = (values[2] - values[1]) / range_

    if Q_min > Q_crit:
        return values[0]
    elif Q_max > Q_crit:
        return values[2]
    else:
        return None

def remove_outliers_fromplate(data):
    samples = data.index.get_level_values('Sample').unique()

    for sample in samples:
        if sample == "NEG":
            continue
        sample_data = data.xs(sample, level="Sample")
        genes = sample_data.index.get_level_values("Gene").unique()
        for gene in genes:
            if gene == "missing_gene":
                continue
            gene_data = sample_data.xs(gene, level="Gene")
            outlier = dixons_q_test_n3(gene_data.values.flatten().tolist())
            if outlier is not None:
                mask = ~((data.index.get_level_values('Sample') == sample) &
                         (data.index.get_level_values('Gene') == gene) &
                         (data['CT'] == outlier))

                data = data[mask]
    return data

#return a table with the mean and standard deviation of each sample's delta CT
def combo_deltact(data):
    samples = data.index.get_level_values('Sample').unique()

    genes = data.index.get_level_values("Gene").unique()
    sampleCol = []
    geneCol = []
    ctMeanCol = []
    SDCol = []

    deltaCTdict = {}

    for sample in samples:
        sample_data = data.xs(sample, level="Sample")
        if sample == "empty":
            continue
        if sample == "NEG":
            continue
        control_cts = sample_data.xs(CONTROL_GENE, level="Gene")
        for gene in genes:
            if gene == CONTROL_GENE:
                continue
            if gene == "missing_gene":
                continue
            exp_cts = sample_data.xs(gene, level="Gene")
            deltaCTs = []
            for CT1 in exp_cts["CT"]:
                for CT2 in control_cts["CT"]:
                    deltaCTs.append( round(CT1 - CT2, 3) )

            sampleCol.append(sample)
            geneCol.append(gene)
            ctMeanCol.append(round(statistics.mean(deltaCTs), 5))
            SDCol.append(round(numpy.std(deltaCTs), 5))
            deltaCTdict[str(sample)] = deltaCTs

    columns = ["Sample", "Gene", "Mean_ΔCT", "STDEV_CT"]
    ctData = zip(sampleCol, geneCol, ctMeanCol, SDCol)

    ctData = pd.DataFrame(ctData, columns=columns)
    print(ctData)
    ctData['Sample'] = ctData['Sample'].astype(str) # gotta force everything as a string otherwise pandas allows a mix

    control_deltact = ctData.loc[ctData['Sample'].str.startswith('C'), 'Mean_ΔCT'].mean()

    deltadeltacts = {
        gene: [round(Δct - control_deltact, 5) for Δct in value_list]
        for gene, value_list in deltaCTdict.items()
    }
    print(deltadeltacts)

    knockdowns = {
        gene: [round(2 ** (-1 * Δ),5) for Δ in Δs]
        for gene, Δs in deltadeltacts.items()
    }

    # Now to calculate 2^-deltadeltaCT
    ctData["Knockdown"] = ctData['Sample'].apply(lambda x: numpy.mean(knockdowns[x]) if x in deltadeltacts else numpy.nan)

    # Add mean column
    ctData['Mean_ΔΔCT'] = ctData['Sample'].apply(lambda x: numpy.mean(deltadeltacts[x]) if x in deltadeltacts else numpy.nan)

    # Add standard deviation column
    ctData['STDEV_KD'] = ctData['Sample'].apply(lambda x: numpy.std(knockdowns[x]) if x in deltadeltacts else numpy.nan)

    # Add percentage knockdown column
    ctData['Percentage'] = ctData['Knockdown'] * 100

    # Calculate upper and lower bounds
    ctData['STDEV_%'] = round((ctData['STDEV_KD'] * 100), 5)
    ctData['LowerBound'] = round((ctData['Knockdown'] - ctData['STDEV_KD']) * 100, 5)
    ctData['UpperBound'] = round((ctData['Knockdown'] + ctData['STDEV_KD']) * 100, 5)

    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 2000)
    return ctData

def savedata_excel(excelfile, analysisdata):
    book = load_workbook(excelfile)

    # Remove 'analysis' sheet if it exists
    if 'analysis' in book.sheetnames:
        std = book.get_sheet_by_name('analysis')
        book.remove(std)

    # Save and close the modified workbook
    book.save(excelfile)
    book.close()

    # Append to the existing file
    with pd.ExcelWriter(excelfile, engine='openpyxl', mode='a') as writer:
        analysisdata.to_excel(writer, sheet_name='analysis', index=False)

    # Reopen to apply formatting
    book = load_workbook(excelfile)
    worksheet = book['analysis']

    # Format the header row
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    numeric_columns = {}

    for col_num, column in enumerate(analysisdata.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill

        # Check for numeric data in column (for later formatting)
        numeric_columns[col_num] = pd.api.types.is_numeric_dtype(analysisdata[column])

        # Adjust column width to fit content
        column_letter = get_column_letter(col_num)
        max_length = max(analysisdata[column].astype(str).map(len).max(), len(str(column)))
        adjusted_width = (max_length + 2) * 1.2  # Add some padding
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Add zebra striping (alternating row colors)
    for row_idx in range(2, len(analysisdata) + 2):
        for col_idx in range(1, len(analysisdata.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)

            # Apply zebra striping
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                # Apply number formatting (5 decimal places) to numeric columns
                if numeric_columns[col_idx]:
                    cell.number_format = '0.#####'  # Custom format for exactly 5 decimal places

    # Save the formatted workbook
    book.save(excelfile)
    book.close()


df_plate = roche_html_extractor(htmlPath)
qpcr_data = apply_design_xlsx(wb_path, df_plate)
# Output the DataFrame to a CSV file
df_plate.to_csv(outfile_csv_name)
qpcr_data_filtered = remove_outliers_fromplate(qpcr_data)
ct_data = combo_deltact(qpcr_data_filtered)


# Output the plate data and the ct data
savedata_excel(wb_path, ct_data)
print(df_plate)
print(ct_data)
